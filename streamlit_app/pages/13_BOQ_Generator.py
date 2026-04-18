"""BOQ Generator — OpenMEP"""

import streamlit as st
import sys
import os
import io
import pandas as pd
from datetime import date

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils import (
    apply_theme_css, RED, BLACK, WHITE, DARK_GREY, LIGHT_GREY,
    page_header, result_card, section_title, api_post, region_selector,
    TEAL, TEAL_L
)

st.set_page_config(page_title="BOQ Generator — OpenMEP", page_icon="📋", layout="wide")
apply_theme_css()

page_header("BOQ Generator", "Generate Bill of Quantities from MEP cable, duct, and pipe take-offs", "📋")

with st.sidebar:
    st.markdown(f"<h3 style='color:{TEAL_L}; font-size:0.9rem; text-transform:uppercase; letter-spacing:2px;'>BOQ Format</h3>", unsafe_allow_html=True)
    st.markdown(f"""
    <div style="background:{DARK_GREY}; padding:1rem; border-radius:6px; font-size:0.8rem; color:{WHITE}; line-height:1.9;">
    API generates priced BOQ from:<br><br>
    <b style='color:{TEAL_L}'>Cables:</b><br>
    circuit ref, type, size, length<br>
    conduit / tray included<br><br>
    <b style='color:{TEAL_L}'>Ductwork:</b><br>
    width × height, length, material<br><br>
    <b style='color:{TEAL_L}'>Pipework:</b><br>
    DN, length, material, system<br><br>
    Rates in USD by default.<br>
    GCC → multiply by 3.67 for AED.
    </div>
    """, unsafe_allow_html=True)

# ── Session state ──────────────────────────────────────────────────────────────
if "boq_cables" not in st.session_state:
    st.session_state.boq_cables = [
        {"circuit_reference": "HV/LV-T01", "description": "Feeder cable — MDB to SMDB-L3", "cable_type": "XLPE_CU", "cable_size_mm2": 95.0, "phases": 3, "cable_length_m": 120.0, "runs": 1, "conduit_diameter_mm": None, "tray_width_mm": 300.0},
        {"circuit_reference": "SMDB-C1", "description": "Lighting circuit C1", "cable_type": "XLPE_CU", "cable_size_mm2": 2.5, "phases": 1, "cable_length_m": 45.0, "runs": 1, "conduit_diameter_mm": 25.0, "tray_width_mm": None},
        {"circuit_reference": "SMDB-C3", "description": "Power sockets circuit", "cable_type": "XLPE_CU", "cable_size_mm2": 4.0, "phases": 1, "cable_length_m": 30.0, "runs": 1, "conduit_diameter_mm": 25.0, "tray_width_mm": None},
        {"circuit_reference": "SMDB-C4", "description": "AHU supply cable", "cable_type": "XLPE_CU", "cable_size_mm2": 16.0, "phases": 3, "cable_length_m": 25.0, "runs": 1, "conduit_diameter_mm": None, "tray_width_mm": 150.0},
    ]

if "boq_ducts" not in st.session_state:
    st.session_state.boq_ducts = [
        {"section_id": "SD-01", "description": "Main supply duct — AHU to VAV", "width_mm": 600.0, "height_mm": 400.0, "diameter_mm": None, "duct_type": "rectangular", "length_m": 40.0, "material": "galvanised_steel"},
        {"section_id": "SD-02", "description": "Branch duct to diffusers", "width_mm": 300.0, "height_mm": 200.0, "diameter_mm": None, "duct_type": "rectangular", "length_m": 20.0, "material": "galvanised_steel"},
    ]

if "boq_pipes" not in st.session_state:
    st.session_state.boq_pipes = [
        {"section_id": "CW-01", "description": "Cold water distribution main", "nominal_dn": 50, "diameter_mm": 50.0, "length_m": 80.0, "material": "copper", "system": "CWDS", "insulation": False},
        {"section_id": "HW-01", "description": "Hot water supply pipe", "nominal_dn": 32, "diameter_mm": 32.0, "length_m": 60.0, "material": "copper", "system": "HWDS", "insulation": True},
    ]

section_title("Project & Regional Settings")
region_code, sub_code = region_selector("boq")
col_t, col_c = st.columns(2)
with col_t:
    project_title = st.text_input("Project Title", value=st.session_state.get("project_name", "MEP Engineering Project"))
with col_c:
    _boq_schema = {
        "gcc": "FIDIC / NEC3 (AED)",
        "europe": "NRM2 / SMM7 (GBP)",
        "india": "CPWD DSR / IS standards (INR)",
        "australia": "AIQS / AISC (AUD)",
    }
    st.info(f"BOQ schema: {_boq_schema.get(region_code, 'FIDIC/USD')} — authority: {sub_code.upper()}")

# ── Cable Items ──────────────────────────────────────────────────────────────────
section_title("Cable Take-Off")
CABLE_TYPE_OPTS = ["XLPE_CU", "PVC_CU", "XLPE_AL", "FRLS_CU", "X90_CU"]
col1, col2, col3, col4, col5, col6, col7 = st.columns([2, 3, 1.5, 1.2, 1, 1.5, 1.5])
for c, h in zip([col1,col2,col3,col4,col5,col6,col7], ["Ref", "Description", "Cable Type", "Size mm²", "Ph", "Length (m)", "Runs"]):
    c.markdown(f"<span style='color:{LIGHT_GREY}; font-size:0.8rem; font-weight:600;'>{h}</span>", unsafe_allow_html=True)

for i, item in enumerate(st.session_state.boq_cables):
    c1, c2, c3, c4, c5, c6, c7 = st.columns([2, 3, 1.5, 1.2, 1, 1.5, 1.5])
    with c1:
        item["circuit_reference"] = st.text_input(f"bc_r_{i}", value=item["circuit_reference"], label_visibility="collapsed", key=f"bc_r_{i}")
    with c2:
        item["description"] = st.text_input(f"bc_d_{i}", value=item["description"], label_visibility="collapsed", key=f"bc_d_{i}")
    with c3:
        ct_idx = CABLE_TYPE_OPTS.index(item["cable_type"]) if item["cable_type"] in CABLE_TYPE_OPTS else 0
        item["cable_type"] = st.selectbox(f"bc_t_{i}", CABLE_TYPE_OPTS, index=ct_idx, label_visibility="collapsed", key=f"bc_t_{i}")
    with c4:
        item["cable_size_mm2"] = st.number_input(f"bc_s_{i}", value=float(item["cable_size_mm2"]), min_value=1.5, max_value=630.0, step=0.5, label_visibility="collapsed", key=f"bc_s_{i}")
    with c5:
        item["phases"] = st.selectbox(f"bc_p_{i}", [1, 3], index=0 if item["phases"] == 1 else 1, label_visibility="collapsed", key=f"bc_p_{i}")
    with c6:
        item["cable_length_m"] = st.number_input(f"bc_l_{i}", value=float(item["cable_length_m"]), min_value=0.0, step=5.0, label_visibility="collapsed", key=f"bc_l_{i}")
    with c7:
        item["runs"] = st.number_input(f"bc_rn_{i}", value=int(item["runs"]), min_value=1, max_value=10, step=1, label_visibility="collapsed", key=f"bc_rn_{i}")

col1, col2 = st.columns([2, 10])
with col1:
    if st.button("+ Add Cable"):
        st.session_state.boq_cables.append({"circuit_reference": f"C{len(st.session_state.boq_cables)+1}", "description": "New Cable", "cable_type": "XLPE_CU", "cable_size_mm2": 2.5, "phases": 3, "cable_length_m": 20.0, "runs": 1, "conduit_diameter_mm": None, "tray_width_mm": None})
        st.rerun()

# ── Duct Items ──────────────────────────────────────────────────────────────────
section_title("Ductwork Take-Off")
DUCT_MATERIALS = ["galvanised_steel", "stainless_steel", "aluminium", "pvc"]
df_ducts = pd.DataFrame(st.session_state.boq_ducts)
st.dataframe(df_ducts, use_container_width=True, hide_index=True)

with st.expander("Edit / Add Ductwork"):
    with st.form("add_duct_form"):
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            d_ref = st.text_input("Section ID", placeholder="SD-03")
            d_desc = st.text_input("Description")
        with col2:
            d_w = st.number_input("Width (mm)", value=400.0, step=50.0)
            d_h = st.number_input("Height (mm)", value=250.0, step=50.0)
        with col3:
            d_len = st.number_input("Length (m)", value=15.0, step=5.0)
            d_mat = st.selectbox("Material", DUCT_MATERIALS)
        with col4:
            d_type = st.selectbox("Type", ["rectangular", "circular", "oval"])
            d_dia = st.number_input("Diameter mm (circular only)", value=0.0, step=25.0)
        if st.form_submit_button("Add Duct Section"):
            st.session_state.boq_ducts.append({"section_id": d_ref, "description": d_desc, "width_mm": d_w, "height_mm": d_h, "diameter_mm": d_dia if d_type == "circular" else None, "duct_type": d_type, "length_m": d_len, "material": d_mat})
            st.rerun()

# ── Pipe Items ──────────────────────────────────────────────────────────────────
section_title("Pipework Take-Off")
PIPE_MATERIALS = ["copper", "steel", "pvc", "cpvc", "hdpe", "stainless_steel"]
PIPE_SYSTEMS = ["CWDS", "HWDS", "CWS_main", "LTHW", "CHW", "condenser", "drainage", "gas"]
df_pipes = pd.DataFrame(st.session_state.boq_pipes)
st.dataframe(df_pipes, use_container_width=True, hide_index=True)

with st.expander("Edit / Add Pipework"):
    with st.form("add_pipe_form"):
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            p_ref = st.text_input("Section ID", placeholder="CW-03")
            p_desc = st.text_input("Description")
        with col2:
            p_dn = st.number_input("Nominal DN (mm)", value=25, step=5)
            p_dia = st.number_input("Diameter (mm)", value=25.0, step=5.0)
        with col3:
            p_len = st.number_input("Length (m)", value=20.0, step=5.0)
            p_mat = st.selectbox("Material", PIPE_MATERIALS)
        with col4:
            p_sys = st.selectbox("System", PIPE_SYSTEMS)
            p_ins = st.checkbox("Insulated?", value=False)
        if st.form_submit_button("Add Pipe Section"):
            st.session_state.boq_pipes.append({"section_id": p_ref, "description": p_desc, "nominal_dn": int(p_dn), "diameter_mm": float(p_dia), "length_m": float(p_len), "material": p_mat, "system": p_sys, "insulation": p_ins})
            st.rerun()

# ── Generate ───────────────────────────────────────────────────────────────────
section_title("Generate BOQ")
if st.button("Generate Priced BOQ", use_container_width=True):
    # API MepBoQRequest: project_name, project_reference, region, currency, cables, ducts, pipes
    cables_payload = []
    for item in st.session_state.boq_cables:
        cables_payload.append({
            "circuit_reference": item["circuit_reference"],
            "description": item["description"],
            "cable_type": item["cable_type"],
            "cable_size_mm2": float(item["cable_size_mm2"]),
            "phases": int(item["phases"]),
            "cable_length_m": float(item["cable_length_m"]),
            "runs": int(item["runs"]),
            "conduit_diameter_mm": item.get("conduit_diameter_mm"),
            "tray_width_mm": item.get("tray_width_mm"),
        })
    ducts_payload = []
    for item in st.session_state.boq_ducts:
        ducts_payload.append({
            "section_id": item["section_id"],
            "description": item["description"],
            "width_mm": float(item["width_mm"]),
            "height_mm": float(item["height_mm"]),
            "diameter_mm": item.get("diameter_mm"),
            "duct_type": item["duct_type"],
            "length_m": float(item["length_m"]),
            "material": item["material"],
        })
    pipes_payload = []
    for item in st.session_state.boq_pipes:
        pipes_payload.append({
            "section_id": item["section_id"],
            "description": item["description"],
            "nominal_dn": int(item["nominal_dn"]),
            "diameter_mm": float(item["diameter_mm"]),
            "length_m": float(item["length_m"]),
            "material": item["material"],
            "system": item["system"],
            "insulation": bool(item["insulation"]),
        })

    payload = {
        "project_name": project_title,
        "project_reference": st.session_state.get("project_number", ""),
        "region": region_code,
        "sub_region": sub_code,
        "currency": "USD",
        "cables": cables_payload,
        "ducts": ducts_payload,
        "pipes": pipes_payload,
    }

    with st.spinner("Generating BOQ..."):
        result = api_post("/api/boq/generate", payload)

    if result:
        section_title("BOQ Results")
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            result_card("Cables (USD)", f"{result.get('cable_total_usd', 0):,.0f}", "USD")
        with col2:
            result_card("Ductwork (USD)", f"{result.get('duct_total_usd', 0):,.0f}", "USD")
        with col3:
            result_card("Pipework (USD)", f"{result.get('pipe_total_usd', 0):,.0f}", "USD")
        with col4:
            result_card("Grand Total (USD)", f"{result.get('grand_total_usd', 0):,.0f}", "USD")

        # Line items table — API returns "line_items" list
        if result.get("line_items"):
            section_title("BOQ Line Items")
            rows = []
            for item in result["line_items"]:
                rows.append({
                    "Discipline": item.get("discipline", ""),
                    "Category": item.get("category", ""),
                    "Description": item.get("description", ""),
                    "Qty": item.get("quantity", ""),
                    "Unit": item.get("unit", ""),
                    "Rate (USD)": f"{item.get('unit_rate_usd', 0):,.2f}" if item.get("unit_rate_usd") is not None else "",
                    "Amount (USD)": f"{item.get('amount_usd', 0):,.0f}" if item.get("amount_usd") is not None else "",
                })
            df = pd.DataFrame(rows)
            st.dataframe(df, use_container_width=True, hide_index=True)

        # Region-specific Excel export with appropriate schema and headers
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            schema = result.get("boq_schema", "FIDIC")
            measurement_std = result.get("measurement_standard", "NRM2")
            local_currency = result.get("local_currency", result.get("currency", "USD"))
            usd_fx_rate = result.get("usd_fx_rate", 1.0)
            grand_total_usd = result.get("grand_total_usd", 0)
            grand_total_local = result.get(
                f"grand_total_{local_currency.lower()}",
                grand_total_usd * usd_fx_rate,
            )

            # Column definitions per regional schema
            lc_sym = {"aed": "AED", "gbp": "£", "inr": "INR", "aud": "AUD"}.get(local_currency.lower(), local_currency)
            if region_code == "gcc":
                col_headers = ["Item Ref", "Description of Works", "Unit", "Qty", f"Rate ({lc_sym})", f"Amount ({lc_sym})", "Amount (USD)", "Remarks"]
                schema_title = f"BILL OF QUANTITIES — {schema} / Civil Defence Standards"
                clause_note = "Measured in accordance with FIDIC Conditions of Contract. Rates to include all labour, materials, plant, and overheads."
            elif region_code == "europe":
                col_headers = ["NRM2 Ref", "Description", "Unit", "Quantity", f"Rate ({lc_sym})", f"Amount ({lc_sym})", "Amount (USD)", "Notes"]
                schema_title = f"BILL OF QUANTITIES — {measurement_std} / RICS NRM2"
                clause_note = "Measured in accordance with RICS New Rules of Measurement (NRM2). All rates inclusive of preliminaries and overheads."
            elif region_code == "india":
                col_headers = ["CPWD DSR Ref", "Description of Item", "Unit", "Qty", f"Rate ({lc_sym})", f"Amount ({lc_sym})", "Amount (USD)", "Specification Ref"]
                schema_title = f"BILL OF QUANTITIES — {schema} / CPWD DSR"
                clause_note = "Measured per CPWD Schedule of Rates / DSR. Rates as per current DSR with applicable state CPWD plinth area factor."
            elif region_code == "australia":
                col_headers = ["AIQS Ref", "Description of Works", "Unit", "Quantity", f"Rate ({lc_sym})", f"Amount ({lc_sym})", "Amount (USD)", "Trade Code"]
                schema_title = f"BILL OF QUANTITIES — {schema} / AIQS Standard"
                clause_note = "Measured in accordance with AIQS Cost Management Manual and AISC standards. Rates in AUD inclusive of GST."
            else:
                col_headers = ["Item", "Description", "Unit", "Qty", "Rate (USD)", "Amount (USD)", "Amount (USD)", "Notes"]
                schema_title = "BILL OF QUANTITIES"
                clause_note = ""

            wb = Workbook()
            ws = wb.active
            ws.title = "BOQ"

            header_fill = PatternFill("solid", fgColor="CC0000")
            sub_fill = PatternFill("solid", fgColor="2D2D2D")
            total_fill = PatternFill("solid", fgColor="1A1A1A")
            white_font = Font(color="FFFFFF", bold=True)
            grey_font = Font(color="AAAAAA")
            thin_border = Border(
                left=Side(style="thin", color="444444"),
                right=Side(style="thin", color="444444"),
                bottom=Side(style="thin", color="444444"),
            )

            # Row 1: Schema title
            ws.merge_cells("A1:H1")
            ws["A1"] = schema_title
            ws["A1"].font = Font(color="FFFFFF", bold=True, size=13)
            ws["A1"].fill = header_fill
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 24

            # Row 2: Project info
            ws.merge_cells("A2:D2")
            ws["A2"] = f"Project: {project_title}  |  Ref: {result.get('project_reference', '')}"
            ws["A2"].font = Font(color="DDDDDD", bold=True)
            ws["A2"].fill = sub_fill
            ws.merge_cells("E2:H2")
            ws["E2"] = f"Region: {region_code.upper()}  |  Standard: {schema}  |  Date: {date.today()}"
            ws["E2"].font = Font(color="DDDDDD")
            ws["E2"].fill = sub_fill
            ws.row_dimensions[2].height = 18

            # Row 3: Clause note
            ws.merge_cells("A3:H3")
            ws["A3"] = clause_note
            ws["A3"].font = Font(color="AAAAAA", italic=True, size=9)
            ws.row_dimensions[3].height = 28

            # Row 4: Column headers
            if isinstance(col_headers[0], tuple):
                col_headers = col_headers[0]
            for ci, hdr in enumerate(col_headers, 1):
                cell = ws.cell(row=4, column=ci, value=hdr)
                cell.font = white_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            ws.row_dimensions[4].height = 20

            # Data rows from API line items
            row_num = 5
            line_items = result.get("line_items", [])
            discipline_groups = {}
            for item in line_items:
                d = item.get("discipline", "General")
                discipline_groups.setdefault(d, []).append(item)

            for discipline, items in discipline_groups.items():
                ws.merge_cells(f"A{row_num}:H{row_num}")
                ws[f"A{row_num}"] = f"  {discipline.upper()}"
                ws[f"A{row_num}"].font = Font(color="FF4444", bold=True)
                ws[f"A{row_num}"].fill = sub_fill
                ws.row_dimensions[row_num].height = 16
                row_num += 1

                for item in items:
                    amt_usd = item.get("amount_usd", 0)
                    rate_usd = item.get("unit_rate_usd", 0)
                    # Use API-provided local-currency fields when available
                    lc = local_currency.lower()
                    amt_local = item.get(f"amount_{lc}", amt_usd * usd_fx_rate)
                    rate_local = rate_usd * usd_fx_rate
                    row_data = [
                        item.get("category", ""),
                        item.get("description", ""),
                        item.get("unit", ""),
                        item.get("quantity", ""),
                        f"{rate_local:.2f}",
                        f"{amt_local:.0f}",
                        f"{amt_usd:.0f}",
                        "",
                    ]
                    for ci, val in enumerate(row_data, 1):
                        cell = ws.cell(row=row_num, column=ci, value=val)
                        cell.border = thin_border
                        cell.alignment = Alignment(vertical="center")
                    row_num += 1

            # Grand total row
            row_num += 1
            ws.merge_cells(f"A{row_num}:E{row_num}")
            ws[f"A{row_num}"] = f"GRAND TOTAL — {schema}"
            ws[f"A{row_num}"].font = Font(color="FFFFFF", bold=True)
            ws[f"A{row_num}"].fill = total_fill
            ws[f"F{row_num}"] = f"{grand_total_local:,.0f}"
            ws[f"F{row_num}"].font = Font(color="FF4444", bold=True)
            ws[f"F{row_num}"].fill = total_fill
            ws[f"G{row_num}"] = f"{grand_total_usd:,.0f}"
            ws[f"G{row_num}"].font = Font(color="FF4444", bold=True)
            ws[f"G{row_num}"].fill = total_fill
            ws.row_dimensions[row_num].height = 22

            # Column widths
            col_widths = [14, 52, 8, 10, 14, 14, 14, 20]
            for ci, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(ci)].width = w

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            st.download_button(
                f"⬇️ Download {schema} BOQ (Excel)",
                data=buf.getvalue(),
                file_name=f"BOQ_{schema}_{region_code.upper()}_{project_title.replace(' ','_')}_{date.today()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            st.caption(f"Region: {region_code.upper()} | Schema: {schema} | Standard: {measurement_std} | Currency: {local_currency} | FX Rate (USD→{local_currency}): {usd_fx_rate:.4f}")
        except ImportError:
            st.info("Install openpyxl for Excel export.")
